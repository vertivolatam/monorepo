#!/usr/bin/env python3
"""build_db.py — Construye una base SQLite local del catálogo de cultivos.

Fuentes (solo lectura):
  - apps/raspberry/config/crops.json  (v2 — 107 cultivos, 6 perfiles, provenance)
  - Estructura de Costos.xlsx, hoja "Modelo Fitotécnico" (openpyxl) — columna
    "Parte Comestible / Edible Portion", la clasificación botánica de la hoja
    de la que se deriva el "Tipo de Cultivo / Harvest Type" (Hoja/Fruto/Raíz…).

Salida: apps/raspberry/config/crops.db con 3 tablas:
  - crops          : metadata + sheet_harvest_type + assigned_profile
  - setpoints      : valores OPERATIVOS resueltos del perfil de cada cultivo
  - setpoint_audit : provenance por valor + mecanismo de rollback (is_active)

MODELO DE PERSISTENCIA — la DB es la FUENTE DE VERDAD:
  crops.db es estado runtime por-device y la fuente de verdad durante la
  experimentación local. crops.json es solo el SEED inicial. Por eso:
    * default  = "seed si está vacía o no existe" (NUNCA borra audit existente)
    * --reseed = reset manual destructivo desde crops.json (DROP + reseed)
  Las ediciones del explorador (cambiar setpoint, rollback) persisten en
  crops.db y escriben nuevos registros en setpoint_audit; esa historia no se
  pierde al re-correr build_db.py.

Uso:
  python build_db.py                # seed-if-empty (idempotente, conserva audit)
  python build_db.py --reseed       # reset destructivo desde crops.json
  python build_db.py --crops X.json --xlsx Y.xlsx --db Z.db
  python build_db.py --smoke        # verificación sin GUI (conteos + discrepancias)
"""

import argparse
import json
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

# --- Auto-ensure openpyxl (la hoja XLSX es opcional pero recomendada) --------
try:
    import openpyxl  # noqa: F401
except ImportError:  # pragma: no cover
    import subprocess

    print("[build_db] openpyxl no encontrado — instalando con pip…")
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "openpyxl"], check=False
    )
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        openpyxl = None  # seguimos sin la hoja; sheet_harvest_type vendrá del json
        print("[build_db] openpyxl no disponible — se omitirá lectura del XLSX.")

# --- Rutas por defecto (este archivo vive en tools/crop_explorer/) -----------
HERE = Path(__file__).resolve().parent
CONFIG_DIR = HERE.parent.parent / "config"
DEFAULT_CROPS = CONFIG_DIR / "crops.json"
DEFAULT_DB = CONFIG_DIR / "crops.db"
DEFAULT_XLSX = CONFIG_DIR / "modelo-fitotecnico.xlsx"
SHEET_NAME = "Modelo Fitotécnico"

# Columnas (1-indexed) de la hoja "Modelo Fitotécnico" (header en fila 4).
COL_AEROPONIC = 1
COL_NAME_ES = 4
COL_EDIBLE_PART = 9  # "Parte Comestible / Edible Portion"
SHEET_HEADER_ROW = 4
SHEET_FIRST_DATA_ROW = 5

# --- Grupo verde "Nutritional value per 100 g (3.5 oz)" (cols 106-131) -------
# Cada entrada: clave estable -> (col 1-indexed, unidad). El orden refleja la
# hoja (energía, macros, vitaminas A→K, minerales Ca→Zn). value_num en estas
# unidades; el explorador agrupa por prefijo (vit_*/min_* → Vitaminas/Minerales).
NUTRITION_COLUMNS: list[tuple[str, int, str]] = [
    ("energy", 106, "kJ"),
    ("carbohydrates_sugars", 107, "g"),
    ("dietary_fiber", 108, "g"),
    ("fat", 109, "g"),
    ("protein", 110, "g"),
    ("water", 111, "g"),
    ("vit_a_retinol", 112, "µg"),
    ("vit_beta_carotene", 113, "µg"),
    ("vit_b1_thiamine", 114, "mg"),
    ("vit_b2_riboflavin", 115, "mg"),
    ("vit_b3_niacin", 116, "mg"),
    ("vit_b5_pantothenic", 117, "mg"),
    ("vit_b6", 118, "mg"),
    ("vit_b9_folate", 119, "µg"),
    ("vit_choline", 120, "mg"),
    ("vit_c", 121, "mg"),
    ("vit_e", 122, "mg"),
    ("vit_k", 123, "µg"),
    ("min_calcium", 124, "mg"),
    ("min_iron", 125, "mg"),
    ("min_magnesium", 126, "mg"),
    ("min_manganese", 127, "mg"),
    ("min_phosphorus", 128, "mg"),
    ("min_potassium", 129, "mg"),
    ("min_sodium", 130, "mg"),
    ("min_zinc", 131, "mg"),
]
# Unidad por nutriente (para poblar la tabla y la UI).
NUTRITION_UNITS: dict[str, str] = {k: u for k, _c, u in NUTRITION_COLUMNS}

# --- Clasificación de harvest type ------------------------------------------
# La hoja registra "Parte Comestible" (su propia clasificación botánica). De ahí
# se deriva el "Tipo de Cultivo / Harvest Type" en categorías comparables con el
# perfil asignado.
EDIBLE_PART_TO_SHEET_TYPE = {
    "Hojas (Leaves)": "Cultivos de Hoja",
    "Tallo (Stem)": "Cultivos de Hoja",
    "Inflorescencia (Flower)": "Cultivos de Hoja",
    "Fruto (Fruit)": "Cultivos de Fruto",
    "Raíz (Roots)": "Cultivos de Raíz",
    "Bulbo (Bulb)": "Cultivos de Raíz",
    "Tubérculo (Tuber)": "Cultivos de Raíz",
    "Rizoma (Rhizome)": "Cultivos de Raíz",
    "Semilla o Granos (Seed or Grain)": "Semillas/Granos",
    "No aplica (Not applicable)": None,
}

# Tipo de hoja que CADA perfil presupone (para detectar discrepancias).
PROFILE_TO_SHEET_TYPE = {
    "leafy_vegetative": "Cultivos de Hoja",
    "herb_aromatic": "Cultivos de Hoja",  # hierbas = parte foliar, pero perfil distinto
    "fruiting_vegetative": "Cultivos de Fruto",
    "fruiting_reproductive": "Cultivos de Fruto",
    "berry_fruiting": "Cultivos de Fruto",
    "root_bulking": "Cultivos de Raíz",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# --- Esquema ----------------------------------------------------------------
# CREATE IF NOT EXISTS: el esquema se asegura sin destruir datos existentes.
# El seeding se gatea aparte (seed-if-empty). Solo --reseed hace DROP.
DROP_SCHEMA = """
DROP TABLE IF EXISTS nutrition;
DROP TABLE IF EXISTS setpoint_audit;
DROP TABLE IF EXISTS setpoints;
DROP TABLE IF EXISTS crops;
"""

SCHEMA = """
CREATE TABLE IF NOT EXISTS crops (
    id                INTEGER PRIMARY KEY,
    name_es           TEXT NOT NULL,
    name_en           TEXT,
    family            TEXT,
    species           TEXT,
    origin            TEXT,
    general_use       TEXT,
    common_use        TEXT,
    edible_part       TEXT,
    sheet_harvest_type TEXT,
    aeroponic         INTEGER NOT NULL DEFAULT 0,
    priority          REAL,
    assigned_profile  TEXT
);

CREATE TABLE IF NOT EXISTS setpoints (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    crop_id    INTEGER NOT NULL REFERENCES crops(id),
    field      TEXT NOT NULL,
    value_num  REAL,
    value_text TEXT
);
CREATE INDEX IF NOT EXISTS idx_setpoints_crop  ON setpoints(crop_id);
CREATE INDEX IF NOT EXISTS idx_setpoints_field ON setpoints(crop_id, field);

CREATE TABLE IF NOT EXISTS setpoint_audit (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    crop_id      INTEGER NOT NULL REFERENCES crops(id),
    field        TEXT NOT NULL,
    value_num    REAL,
    value_text   TEXT,
    source       TEXT,           -- sheet | researched | experiment | agronomist
    confidence   TEXT,
    citation     TEXT,
    note         TEXT,
    is_active    INTEGER NOT NULL DEFAULT 1,
    supersedes_id INTEGER REFERENCES setpoint_audit(id),
    changed_at   TEXT NOT NULL,
    changed_by   TEXT
);
CREATE INDEX IF NOT EXISTS idx_audit_crop        ON setpoint_audit(crop_id);
CREATE INDEX IF NOT EXISTS idx_audit_crop_field  ON setpoint_audit(crop_id, field);
CREATE INDEX IF NOT EXISTS idx_audit_active      ON setpoint_audit(crop_id, field, is_active);

CREATE TABLE IF NOT EXISTS nutrition (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    crop_id   INTEGER NOT NULL REFERENCES crops(id),
    nutrient  TEXT NOT NULL,        -- clave estable (energy, protein, vit_c, min_iron…)
    value_num REAL,
    unit      TEXT,
    source    TEXT                  -- sheet | researched
);
CREATE INDEX IF NOT EXISTS idx_nutrition_crop ON nutrition(crop_id);
"""


# --- Lectura del XLSX -------------------------------------------------------
def read_sheet_harvest_types(xlsx_path: Path) -> dict[str, str]:
    """Devuelve {name_es_lower: sheet_harvest_type} derivado de la columna
    'Parte Comestible' de la hoja. Vacío si no hay openpyxl o no existe el archivo."""
    if openpyxl is None or not xlsx_path.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if SHEET_NAME not in wb.sheetnames:
            return {}
        ws = wb[SHEET_NAME]
        out: dict[str, str] = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            if len(row) < COL_EDIBLE_PART:
                continue
            name = row[COL_NAME_ES - 1]
            edible = row[COL_EDIBLE_PART - 1]
            if not name:
                continue
            sheet_type = EDIBLE_PART_TO_SHEET_TYPE.get(
                (edible or "").strip() if isinstance(edible, str) else edible
            )
            out[str(name).strip().lower()] = sheet_type
        return out
    except Exception as exc:  # pragma: no cover
        print(f"[build_db] aviso: no se pudo leer XLSX ({exc}); se usará el json.")
        return {}


def read_sheet_nutrition(xlsx_path: Path) -> dict[int, list[dict]]:
    """Lee el grupo verde 'Nutritional value per 100 g' (cols 106-131) de la hoja,
    indexado por número de fila del XLSX (1-indexed). Cada cultivo del json trae
    su ``source_row``; con él recuperamos sus celdas nutricionales.

    Devuelve {source_row: [{nutrient, value_num, unit, source='sheet'}, …]}.
    Solo incluye celdas numéricas presentes. Vacío si no hay openpyxl/archivo.

    NOTA: al 2026-06-21 este grupo está VACÍO en la hoja (0 celdas en las 1081
    filas), igual que las fases Reproductiva/Maduración. El extractor queda listo
    para cuando se pueble; mientras tanto la nutrición visible viene del bloque
    ``nutrition_per_100g`` (source='researched') de crops.json."""
    if openpyxl is None or not xlsx_path.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if SHEET_NAME not in wb.sheetnames:
            return {}
        ws = wb[SHEET_NAME]
        out: dict[int, list[dict]] = {}
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=SHEET_FIRST_DATA_ROW, values_only=True),
            start=SHEET_FIRST_DATA_ROW,
        ):
            entries: list[dict] = []
            for nutrient, col, unit in NUTRITION_COLUMNS:
                if len(row) < col:
                    continue
                val = row[col - 1]
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    entries.append(
                        {
                            "nutrient": nutrient,
                            "value_num": float(val),
                            "unit": unit,
                            "source": "sheet",
                        }
                    )
            if entries:
                out[r_idx] = entries
        return out
    except Exception as exc:  # pragma: no cover
        print(f"[build_db] aviso: no se pudo leer nutrición del XLSX ({exc}).")
        return {}


def resolve_nutrition(crop: dict, sheet_nutrition: dict[int, list[dict]]) -> list[dict]:
    """Resuelve las filas de nutrición de un cultivo. Prioriza el grupo verde de
    la hoja (vía ``source_row``); si la hoja no tiene datos para ese cultivo,
    usa el bloque ``nutrition_per_100g`` de crops.json (source='researched').

    Cada item: {nutrient, value_num, unit, source}."""
    src_row = crop.get("source_row")
    if src_row is not None and src_row in sheet_nutrition:
        return list(sheet_nutrition[src_row])

    block = crop.get("nutrition_per_100g")
    if not isinstance(block, dict):
        return []
    # Claves meta del bloque que NO son nutrientes.
    _META_KEYS = {"source", "citation", "confidence", "note"}
    rows: list[dict] = []
    for nutrient, payload in block.items():
        if nutrient in _META_KEYS or nutrient not in NUTRITION_UNITS:
            continue
        if isinstance(payload, dict):
            value = payload.get("value")
            unit = payload.get("unit") or NUTRITION_UNITS.get(nutrient)
            source = payload.get("source") or block.get("source") or "researched"
        else:
            value = payload
            unit = NUTRITION_UNITS.get(nutrient)
            source = block.get("source") or "researched"
        if value is None:
            continue
        rows.append(
            {
                "nutrient": nutrient,
                "value_num": float(value),
                "unit": unit,
                "source": source,
            }
        )
    return rows


def derive_sheet_type(crop: dict, sheet_lookup: dict[str, str]) -> str | None:
    """Prefiere el valor de la hoja; si falta, lo deriva de edible_part del json."""
    name = (crop.get("name_es") or "").strip().lower()
    if name in sheet_lookup and sheet_lookup[name] is not None:
        return sheet_lookup[name]
    edible = crop.get("edible_part")
    if isinstance(edible, str):
        return EDIBLE_PART_TO_SHEET_TYPE.get(edible.strip())
    return None


# --- Resolución de setpoints operativos desde el perfil ---------------------
def _prov(d: dict, default_source: str = "researched") -> dict:
    """Extrae provenance común de un sub-dict del perfil."""
    return {
        "source": d.get("source", default_source),
        "confidence": d.get("confidence"),
        "citation": d.get("citation"),
        "note": d.get("research_note") or d.get("note"),
    }


def resolve_setpoints(profile: dict) -> list[dict]:
    """Desenvuelve la provenance del perfil en una lista plana de setpoints
    operativos. Cada item: {field, value_num?, value_text?, source, confidence,
    citation, note}."""
    rows: list[dict] = []

    def emit(field, *, num=None, text=None, prov=None):
        p = prov or {}
        rows.append(
            {
                "field": field,
                "value_num": num,
                "value_text": text,
                "source": p.get("source"),
                "confidence": p.get("confidence"),
                "citation": p.get("citation"),
                "note": p.get("note"),
            }
        )

    # pH
    ph = profile.get("ph")
    if isinstance(ph, dict):
        p = _prov(ph, "sheet")
        if ph.get("ideal") is not None:
            emit("ph_ideal", num=ph["ideal"], prov=p)
        if ph.get("min") is not None:
            emit("ph_min", num=ph["min"], prov=p)
        if ph.get("max") is not None:
            emit("ph_max", num=ph["max"], prov=p)

    # EC (preferimos dS/m investigado; es el valor operativo real)
    ec = profile.get("ec_dS_m")
    if isinstance(ec, dict):
        p = _prov(ec)
        if ec.get("min") is not None:
            emit("ec_min_dsm", num=ec["min"], prov=p)
        if ec.get("max") is not None:
            emit("ec_max_dsm", num=ec["max"], prov=p)

    # Temperatura ambiente día/noche
    temp = profile.get("ambient_temp_c")
    if isinstance(temp, dict):
        p = _prov(temp, "sheet")
        day = temp.get("day", {})
        night = temp.get("night", {})
        if isinstance(day, dict):
            if day.get("min") is not None:
                emit("temp_day_min", num=day["min"], prov=p)
            if day.get("max") is not None:
                emit("temp_day_max", num=day["max"], prov=p)
        if isinstance(night, dict):
            if night.get("min") is not None:
                emit("temp_night_min", num=night["min"], prov=p)
            if night.get("max") is not None:
                emit("temp_night_max", num=night["max"], prov=p)

    # Humedad relativa (usamos rango día como humidity_min/max operativo)
    rh = profile.get("relative_humidity_pct")
    if isinstance(rh, dict):
        p = _prov(rh, "sheet")
        day = rh.get("day", {})
        if isinstance(day, dict):
            if day.get("min") is not None:
                emit("humidity_min", num=day["min"], prov=p)
            if day.get("max") is not None:
                emit("humidity_max", num=day["max"], prov=p)

    # PPFD
    ppfd = profile.get("ppfd_umol_m2_s")
    if isinstance(ppfd, dict):
        p = _prov(ppfd, "sheet")
        if ppfd.get("min") is not None:
            emit("ppfd_min", num=ppfd["min"], prov=p)
        if ppfd.get("max") is not None:
            emit("ppfd_max", num=ppfd["max"], prov=p)

    # ORP
    orp = profile.get("orp_mv")
    if isinstance(orp, dict):
        p = _prov(orp, "sheet")
        if orp.get("min") is not None:
            emit("orp_min", num=orp["min"], prov=p)
        if orp.get("max") is not None:
            emit("orp_max", num=orp["max"], prov=p)

    # Fotoperiodo
    photo = profile.get("photoperiod_h")
    if isinstance(photo, dict) and photo.get("value") is not None:
        emit("photoperiod_h", num=photo["value"], prov=_prov(photo, "sheet"))

    # Espectro de luz (texto)
    spec = profile.get("light_spectrum")
    if isinstance(spec, dict) and spec.get("value") is not None:
        emit("light_spectrum", text=str(spec["value"]), prov=_prov(spec, "sheet"))

    # Receta de nutrientes (JSON serializado como texto)
    recipe = profile.get("nutrient_recipe_g_per_1000ml")
    if isinstance(recipe, dict):
        src = recipe.get("source", "sheet")
        body = {k: v for k, v in recipe.items() if k != "source"}
        emit(
            "nutrient_recipe_json",
            text=json.dumps(body, ensure_ascii=False),
            prov={"source": "sheet" if "sheet" in str(src) else src, "note": str(src)},
        )

    return rows


# --- Construcción de la DB --------------------------------------------------
def build(crops_path: Path, xlsx_path: Path, db_path: Path, reseed: bool = False) -> dict:
    """Seed-once: crea el esquema si falta y siembra desde crops.json SOLO si la
    tabla crops está vacía. Con reseed=True hace un reset destructivo (DROP +
    reseed), descartando el setpoint_audit de experimentación.

    Devuelve stats; 'seeded' indica si se sembró en esta corrida."""
    data = json.loads(crops_path.read_text(encoding="utf-8"))
    profiles = data.get("profiles", {})
    crops = data.get("crops", [])
    sheet_lookup = read_sheet_harvest_types(xlsx_path)
    sheet_nutrition = read_sheet_nutrition(xlsx_path)

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    if reseed:
        conn.executescript(DROP_SCHEMA)
    conn.executescript(SCHEMA)

    existing = cur.execute("SELECT COUNT(*) FROM crops").fetchone()[0]
    if existing > 0 and not reseed:
        # DB ya poblada — fuente de verdad. No tocamos nada (preserva audit).
        n_sp = cur.execute("SELECT COUNT(*) FROM setpoints").fetchone()[0]
        n_audit = cur.execute("SELECT COUNT(*) FROM setpoint_audit").fetchone()[0]
        n_nut = cur.execute("SELECT COUNT(*) FROM nutrition").fetchone()[0]
        disc_rows = cur.execute(
            "SELECT assigned_profile, sheet_harvest_type FROM crops "
            "WHERE assigned_profile IS NOT NULL AND sheet_harvest_type IS NOT NULL"
        ).fetchall()
        disc = sum(
            1
            for p, s in disc_rows
            if PROFILE_TO_SHEET_TYPE.get(p) and PROFILE_TO_SHEET_TYPE[p] != s
        )
        conn.close()
        return {
            "crops": existing,
            "setpoints": n_sp,
            "audit": n_audit,
            "nutrition": n_nut,
            "discrepancies": disc,
            "sheet_rows": len(sheet_lookup),
            "seeded": False,
        }

    ts = now_iso()
    n_setpoints = 0
    n_audit = 0
    n_nutrition = 0
    discrepancies = 0

    for idx, crop in enumerate(crops, start=1):
        sheet_type = derive_sheet_type(crop, sheet_lookup)
        assigned = crop.get("profile")
        cur.execute(
            "INSERT INTO crops (id, name_es, name_en, family, species, origin, "
            "general_use, common_use, edible_part, sheet_harvest_type, aeroponic, "
            "priority, assigned_profile) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                idx,
                crop.get("name_es"),
                crop.get("name_en"),
                crop.get("family"),
                crop.get("species"),
                crop.get("origin"),
                crop.get("general_use"),
                crop.get("common_use"),
                crop.get("edible_part"),
                sheet_type,
                1 if crop.get("aeroponic") else 0,
                crop.get("priority"),
                assigned,
            ),
        )

        # Detección de discrepancia sheet vs profile
        if assigned and sheet_type:
            expected = PROFILE_TO_SHEET_TYPE.get(assigned)
            if expected and expected != sheet_type:
                discrepancies += 1

        # Nutrición por 100 g (independiente del perfil: aplica a TODO cultivo
        # con datos, incluidos los no-aeropónicos sin perfil asignado).
        for nut in resolve_nutrition(crop, sheet_nutrition):
            cur.execute(
                "INSERT INTO nutrition (crop_id, nutrient, value_num, unit, source) "
                "VALUES (?,?,?,?,?)",
                (idx, nut["nutrient"], nut["value_num"], nut["unit"], nut["source"]),
            )
            n_nutrition += 1

        # Setpoints operativos (solo si el cultivo tiene perfil asignado)
        profile = profiles.get(assigned) if assigned else None
        if not isinstance(profile, dict):
            continue

        for sp in resolve_setpoints(profile):
            cur.execute(
                "INSERT INTO setpoints (crop_id, field, value_num, value_text) "
                "VALUES (?,?,?,?)",
                (idx, sp["field"], sp["value_num"], sp["value_text"]),
            )
            n_setpoints += 1
            cur.execute(
                "INSERT INTO setpoint_audit (crop_id, field, value_num, value_text, "
                "source, confidence, citation, note, is_active, supersedes_id, "
                "changed_at, changed_by) VALUES (?,?,?,?,?,?,?,?,1,NULL,?,?)",
                (
                    idx,
                    sp["field"],
                    sp["value_num"],
                    sp["value_text"],
                    sp["source"],
                    sp["confidence"],
                    sp["citation"],
                    sp["note"],
                    ts,
                    "build_db.py",
                ),
            )
            n_audit += 1

    conn.commit()
    conn.close()

    return {
        "crops": len(crops),
        "setpoints": n_setpoints,
        "audit": n_audit,
        "nutrition": n_nutrition,
        "discrepancies": discrepancies,
        "sheet_rows": len(sheet_lookup),
        "seeded": True,
    }


# --- Smoke test (sin GUI) ---------------------------------------------------
def smoke(db_path: Path) -> int:
    if not db_path.exists():
        print(f"[smoke] FALLO: no existe {db_path}")
        return 1
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    n_crops = cur.execute("SELECT COUNT(*) FROM crops").fetchone()[0]
    n_sp = cur.execute("SELECT COUNT(*) FROM setpoints").fetchone()[0]
    n_audit = cur.execute("SELECT COUNT(*) FROM setpoint_audit").fetchone()[0]
    n_nut = cur.execute("SELECT COUNT(*) FROM nutrition").fetchone()[0]

    # discrepancia: al menos un cultivo donde sheet_harvest_type no concuerda
    # con el tipo presupuesto por su perfil.
    rows = cur.execute(
        "SELECT name_es, sheet_harvest_type, assigned_profile FROM crops "
        "WHERE assigned_profile IS NOT NULL AND sheet_harvest_type IS NOT NULL"
    ).fetchall()
    disc = [
        (n, s, p)
        for (n, s, p) in rows
        if PROFILE_TO_SHEET_TYPE.get(p) and PROFILE_TO_SHEET_TYPE[p] != s
    ]
    conn.close()

    ok = True
    print(f"[smoke] crops      = {n_crops}  (esperado 107)")
    print(f"[smoke] setpoints  = {n_sp}     (esperado > 0)")
    print(f"[smoke] audit      = {n_audit}  (esperado > 0)")
    print(f"[smoke] nutrition  = {n_nut}    (esperado > 0)")
    print(f"[smoke] discrepancias sheet/profile = {len(disc)}")
    if disc[:5]:
        print("[smoke] ejemplos:")
        for n, s, p in disc[:5]:
            print(f"          - {n}: hoja='{s}' vs profile='{p}'")

    if n_crops != 107:
        print("[smoke] FALLO: crops != 107")
        ok = False
    if n_sp <= 0:
        print("[smoke] FALLO: setpoints == 0")
        ok = False
    if n_audit <= 0:
        print("[smoke] FALLO: audit == 0")
        ok = False
    if n_nut <= 0:
        print("[smoke] FALLO: nutrition == 0")
        ok = False
    if len(disc) < 1:
        print("[smoke] FALLO: no se detectó ninguna discrepancia sheet/profile")
        ok = False

    print("[smoke] OK" if ok else "[smoke] FALLÓ")
    return 0 if ok else 1


def main() -> int:
    ap = argparse.ArgumentParser(description="Construye crops.db desde crops.json + XLSX")
    ap.add_argument("--crops", type=Path, default=DEFAULT_CROPS)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--db", type=Path, default=DEFAULT_DB)
    ap.add_argument("--smoke", action="store_true", help="solo correr smoke test")
    ap.add_argument(
        "--reseed",
        "--force",
        action="store_true",
        dest="reseed",
        help="RESET destructivo desde crops.json (descarta setpoint_audit existente)",
    )
    args = ap.parse_args()

    if args.smoke:
        return smoke(args.db)

    if not args.crops.exists():
        print(f"[build_db] ERROR: no existe {args.crops}")
        return 1

    print(f"[build_db] crops.json : {args.crops}")
    print(f"[build_db] xlsx       : {args.xlsx} ({'ok' if args.xlsx.exists() else 'ausente'})")
    print(f"[build_db] db salida  : {args.db}")
    if args.reseed:
        print("[build_db] modo --reseed: RESET destructivo desde crops.json")
    stats = build(args.crops, args.xlsx, args.db, reseed=args.reseed)
    if stats["seeded"]:
        print(
            f"[build_db] sembrada — crops={stats['crops']} setpoints={stats['setpoints']} "
            f"audit={stats['audit']} nutrition={stats['nutrition']} "
            f"discrepancias={stats['discrepancies']} (filas-hoja={stats['sheet_rows']})"
        )
    else:
        print(
            f"[build_db] DB ya poblada (fuente de verdad) — no se reseed. "
            f"crops={stats['crops']} setpoints={stats['setpoints']} "
            f"audit={stats['audit']} discrepancias={stats['discrepancies']}. "
            f"Usa --reseed para reset destructivo."
        )
    return smoke(args.db)


if __name__ == "__main__":
    sys.exit(main())
